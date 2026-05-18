import csv
import io
from datetime import datetime, time, timezone

from flask import g, request, send_file, current_app
from marshmallow import ValidationError

from . import admin_bp
from ...extensions import db
from ...models import Stand, EventDate, Shift, Volunteer, Registration
from ...utils.auth import require_staff
from ...utils.responses import ok, error


_IMPORT_COLUMNS = ['Stand', 'Datum (TT.MM.JJJJ)', 'Von (HH:MM)', 'Bis (HH:MM)', 'Max. Helfer']


# ---------------------------------------------------------------------------
# Template-Download
# ---------------------------------------------------------------------------

@admin_bp.route('/<slug>/import/template/csv', methods=['GET'])
@require_staff
def import_template_csv(slug):
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(_IMPORT_COLUMNS)
    writer.writerow(['Kasse', '01.08.2025', '08:00', '12:00', '3'])
    output.seek(0)
    buf = io.BytesIO(('﻿' + output.getvalue()).encode('utf-8'))
    return send_file(buf, mimetype='text/csv', as_attachment=True,
                     download_name='schicht-vorlage.csv')


@admin_bp.route('/<slug>/import/template/ods', methods=['GET'])
@require_staff
def import_template_ods(slug):
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
    except ImportError:
        return error('odfpy nicht installiert', 500)

    doc = OpenDocumentSpreadsheet()
    table = Table(name='Schichten')
    for row_data in [_IMPORT_COLUMNS, ['Kasse', '01.08.2025', '08:00', '12:00', '3']]:
        row = TableRow()
        for val in row_data:
            cell = TableCell(valuetype='string')
            cell.addElement(P(text=str(val)))
            row.addElement(cell)
        table.addElement(row)
    doc.spreadsheet.addElement(table)

    buf = io.BytesIO()
    doc.write(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype='application/vnd.oasis.opendocument.spreadsheet',
                     as_attachment=True, download_name='schicht-vorlage.ods')


@admin_bp.route('/<slug>/import/template/xlsx', methods=['GET'])
@require_staff
def import_template_xlsx(slug):
    try:
        import openpyxl
    except ImportError:
        return error('openpyxl nicht installiert', 500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Schichten'
    ws.append(_IMPORT_COLUMNS)
    ws.append(['Kasse', '01.08.2025', '08:00', '12:00', 3])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='schicht-vorlage.xlsx')


# ---------------------------------------------------------------------------
# Import-Endpunkte
# ---------------------------------------------------------------------------

@admin_bp.route('/<slug>/import/shifts/csv', methods=['POST'])
@require_staff
def import_shifts_csv(slug):
    if 'file' not in request.files:
        return error('Keine Datei übergeben', 400)
    content = request.files['file'].read().decode('utf-8-sig')
    rows = list(csv.DictReader(io.StringIO(content), delimiter=';'))
    return _process_shift_rows(rows, g.instance.id)


@admin_bp.route('/<slug>/import/shifts/xlsx', methods=['POST'])
@require_staff
def import_shifts_xlsx(slug):
    if 'file' not in request.files:
        return error('Keine Datei übergeben', 400)
    try:
        import openpyxl
    except ImportError:
        return error('openpyxl nicht installiert', 500)

    wb = openpyxl.load_workbook(request.files['file'])
    ws = wb.active
    header = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
    rows = [
        dict(zip(header, [str(c.value or '').strip() for c in row]))
        for row in ws.iter_rows(min_row=2)
    ]
    return _process_shift_rows(rows, g.instance.id)


@admin_bp.route('/<slug>/import/shifts/ods', methods=['POST'])
@require_staff
def import_shifts_ods(slug):
    if 'file' not in request.files:
        return error('Keine Datei übergeben', 400)
    try:
        from odf.opendocument import load as odf_load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
    except ImportError:
        return error('odfpy nicht installiert', 500)

    doc = odf_load(request.files['file'])
    sheet = doc.spreadsheet.getElementsByType(Table)[0]
    raw_rows = sheet.getElementsByType(TableRow)

    header = None
    rows = []
    for row in raw_rows:
        cells = row.getElementsByType(TableCell)
        values = []
        for cell in cells:
            texts = cell.getElementsByType(P)
            values.append(str(texts[0]) if texts else '')
        if header is None:
            header = [v.strip() for v in values]
        else:
            rows.append(dict(zip(header, [v.strip() for v in values])))

    return _process_shift_rows(rows, g.instance.id)


# ---------------------------------------------------------------------------
# Gemeinsame Verarbeitungslogik
# ---------------------------------------------------------------------------

def _process_shift_rows(rows, instance_id):
    created = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            stand_name = row.get('Stand', '').strip()
            date_str = row.get('Datum (TT.MM.JJJJ)', '').strip()
            start_str = row.get('Von (HH:MM)', '').strip()
            end_str = row.get('Bis (HH:MM)', '').strip()
            max_str = str(row.get('Max. Helfer', '2')).strip()

            if not any([stand_name, date_str, start_str, end_str]):
                continue

            event_date_obj = datetime.strptime(date_str, '%d.%m.%Y').date()
            start_time = time.fromisoformat(start_str)
            end_time = time.fromisoformat(end_str)
            max_volunteers = int(max_str) if max_str.isdigit() else 2

            stand = Stand.query.filter_by(instance_id=instance_id, name=stand_name).first()
            if not stand:
                stand = Stand(instance_id=instance_id, name=stand_name)
                db.session.add(stand)
                db.session.flush()

            event_date = EventDate.query.filter_by(
                instance_id=instance_id, date=event_date_obj
            ).first()
            if not event_date:
                event_date = EventDate(instance_id=instance_id, date=event_date_obj)
                db.session.add(event_date)
                db.session.flush()

            existing = Shift.query.filter_by(
                stand_id=stand.id,
                event_date_id=event_date.id,
                start_time=start_time,
                end_time=end_time,
            ).first()

            if existing:
                skipped += 1
            else:
                db.session.add(Shift(
                    stand_id=stand.id,
                    event_date_id=event_date.id,
                    start_time=start_time,
                    end_time=end_time,
                    max_volunteers=max_volunteers,
                ))
                created += 1

        except Exception as e:
            errors.append(f'Zeile {i}: {e}')

    db.session.commit()
    return ok({'created': created, 'skipped': skipped, 'errors': errors},
              f'{created} Schichten importiert, {skipped} übersprungen')
